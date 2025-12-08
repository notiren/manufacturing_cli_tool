import sys
import subprocess
import os

# ---------------------------------------------------------
# AUTO-INSTALL REQUIRED PACKAGES
# ---------------------------------------------------------
def ensure_package(pkg, imp=None):
    try:
        __import__(imp or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", pkg])

ensure_package("openpyxl")
ensure_package("pandas")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ---------------------------------------------------------
# CONFIG
# ---------------------------------------------------------
HEADER_ROW = 14
LIMIT_ROW = 2
TOLERANCE = 0.3
VENDOR1 = "UNISON"
VENDOR2 = "LCE"

VENDOR1_OUTPUT_COLS = [
    "NO", "SN", "LENS SN", "RESULT", "CT",
    "FOV0.300_LT", "FOV0.300_RT", "FOV0.300_LB", "FOV0.300_RB", "FOV0.300_R",
    "FOV0.650_L", "FOV0.650_R",
    "FOV0.700_LT", "FOV0.700_RT", "FOV0.700_LB", "FOV0.700_RB",
    "FOV0.750_LT", "FOV0.750_RT", "FOV0.750_LB", "FOV0.750_RB"
]

VENDOR2_OUTPUT_COLS = [
    "NO", "SN", "LENS SN", "RESULT",
    "B0_H_V",
    "B1_H_V", "B2_H_V", "B3_H_V", "B4_H_V",
    "B5_H_V", "B6_H_V", "B7_H_V", "B8_H_V", "B9_H_V", "B10_H_V",
    "B11_H_V", "B12_H_V", "B13_H_V", "B14_H_V", "B15_H_V", "B16_H_V"
]

LIMIT_MAP = {
    "1": {
        "CT": "$A$2",
        "FOV0.300": "$B$2",
        "FOV0.650": "$C$2",
        "FOV0.700": "$D$2",
        "FOV0.750": "$E$2",
        "TOL": "$F$2"
    },
    "2": {
        "B0": "$A$2",
        "B1": "$B$2", "B2": "$B$2", "B3": "$B$2", "B4": "$B$2",
        "B5": "$C$2", "B6": "$C$2", "B7": "$C$2", "B8": "$C$2", "B9": "$C$2", "B10": "$C$2",
        "B11": "$D$2", "B12": "$D$2", "B13": "$D$2", "B14": "$D$2", "B15": "$D$2", "B16": "$D$2",
        "TOL": "$E$2"
    }
}

# ---------------------------------------------------------
# VENDOR DATA PROCESSING
# ---------------------------------------------------------
def process_vendor1(df):
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()
    lens_col = "LENS SN"
    out = pd.DataFrame(columns=VENDOR1_OUTPUT_COLS)
    for i, row in df.iterrows():
        new_row = {
            "NO": i+1,
            "SN": row.get("SN", ""),
            "RESULT": "",
            "LENS SN": str(row.get(lens_col, ""))
        }
        for col in VENDOR1_OUTPUT_COLS[4:]:
            new_row[col] = row.get(col, None)
        out.loc[len(out)] = new_row
    return out

def process_vendor2(df):
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()
    sn_col, lens_col, sensorid_col = "SN", "LENS SN", "SensorID"
    out = pd.DataFrame(columns=VENDOR2_OUTPUT_COLS)
    for i, row in df.iterrows():
        new_row = {
            "NO": row.get("NO", i+1),
            "SN": row.get(sn_col, ""),
            "RESULT": "",
            "LENS SN": str(row.get(lens_col, row.get(sensorid_col, "")))
        }
        for idx in range(17):
            h, v = row.get(f"B{idx}_H", None), row.get(f"B{idx}_V", None)
            try: avg = (float(h) + float(v)) / 2
            except: avg = None
            new_row[f"B{idx}_H_V"] = avg
        out.loc[len(out)] = new_row
    return out

# ---------------------------------------------------------
# WIDTH REGULATION HELPER
# ---------------------------------------------------------
def set_column_width(ws, col_idx, min_width=8, max_width=17):
    max_len = 0
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, col_idx).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_width), max_width)

# ---------------------------------------------------------
# FIX LENS SN FORMAT
# ---------------------------------------------------------
def fix_lens_sn_format(ws):
    lens_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(HEADER_ROW, c).value).strip().upper() == "LENS SN":
            lens_col = c
            break

    if not lens_col:
        print("LENS SN column not found")
        return
    
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        cell = ws.cell(r, lens_col)
        val = cell.value
        if val is None:
            continue

        s = str(val).strip()
        if s == "":
            continue

        cell.number_format = "@"  

# ---------------------------------------------------------
# INSERT LIMITS TABLE
# ---------------------------------------------------------
def insert_limits_table(ws, vendor):
    ws.insert_rows(1, amount=HEADER_ROW-1)
    
    gray1 = PatternFill("solid", fgColor="EDEDED")
    gray2 = PatternFill("solid", fgColor="D9D9D9")
    gray3 = PatternFill("solid", fgColor="C4C4C4")
    gray4 = PatternFill("solid", fgColor="AFAFAF")
    tol_font = Font(color="00B0F0", bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    if vendor=="1":
        headers = ["CT (Center)", "(0.3FoV)", "(0.65FoV)", "(0.7FoV)", "(0.75FoV)", "TOLERANCE"]
        values  = [64, 59.7, 47.5, 47.5, 45.1, TOLERANCE]
        fills   = [gray1, gray2, gray3, gray3, gray4, gray1]
        fill_map = [1, 5, 6, 4]
        tol_idx = 6
    else:
        headers = ["B0 (CT)", "B1-B4 (0.3FoV)", "B5-B10 (0.65FoV)", "B11-B16 (0.7FoV)", "TOLERANCE"]
        values  = [64, 45.1, 47.5, 59.7, TOLERANCE]
        fills   = [gray1, gray4, gray3, gray2, gray1]
        fill_map = [1, 4, 6, 6]
        tol_idx = 5

    # Apply limits table + fills
    for col_idx,(h,fill) in enumerate(zip(headers,fills),start=1):
        c = ws.cell(1,col_idx,h)
        c.font = tol_font if col_idx==tol_idx else Font(bold=True)
        c.fill = fill
        c.border = border
        set_column_width(ws, col_idx)

    for col_idx,(v,fill) in enumerate(zip(values,fills),start=1):
        c = ws.cell(2,col_idx,v)
        c.border = border
        set_column_width(ws, col_idx, min_width=8, max_width=17)

    # Apply fills to main header row using fill_map
    col_idx = 5
    for fill, span in zip(fills, fill_map):
        for i in range(span):
            c = ws.cell(HEADER_ROW, col_idx)
            c.fill = fill
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            set_column_width(ws, col_idx)
            col_idx += 1

    # Center everything
    alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row, col).alignment = alignment
    
    ws.freeze_panes = f"A{HEADER_ROW+1}"


# ---------------------------------------------------------
# INSERT RESULT AND SUMMARY
# ---------------------------------------------------------
def insert_excel_formulas(ws, vendor):
    header_row = HEADER_ROW
    # Find RESULT column
    result_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value).upper() == "RESULT":
            result_col = c
            break
    if not result_col:
        return

    # Identify measurement columns
    meas_cols = []
    limits = []
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(header_row, c).value).upper()
        if name and name not in ("NO", "SN", "LENS SN", "RESULT"):
            meas_cols.append(c)
            # Map to limit cell
            limit_cell = None
            for key, cell in LIMIT_MAP[vendor].items():
                if key != "TOL" and name.startswith(key):
                    limit_cell = cell
                    break
            limits.append(limit_cell)

    tol_cell = LIMIT_MAP[vendor]["TOL"]
    helper_start_col = ws.max_column + 1
    helper_cols = []

    # Create hidden helper columns
    for idx, (col, limit_cell) in enumerate(zip(meas_cols, limits)):
        helper_col = helper_start_col + idx
        helper_cols.append(helper_col)
        ws.cell(header_row, helper_col, f"_H{idx}")
        ws.column_dimensions[get_column_letter(helper_col)].hidden = True
        for r in range(header_row + 1, ws.max_row + 1):
            formula = (f'=IF({get_column_letter(col)}{r}="", "", '
                       f'IF({get_column_letter(col)}{r} < ({limit_cell}-{tol_cell}),1, '
                       f'IF({get_column_letter(col)}{r} < ({limit_cell}+{tol_cell}),2,3)))')
            ws.cell(r, helper_col).value = formula

    # RESULT column formulas
    first_helper_letter = get_column_letter(helper_cols[0])
    last_helper_letter = get_column_letter(helper_cols[-1])
    for r in range(header_row + 1, ws.max_row + 1):
        ws.cell(r, result_col).value = (
            f'=IF(COUNTIF({first_helper_letter}{r}:{last_helper_letter}{r},1)>0,"FAIL",'
            f'IF(COUNTIF({first_helper_letter}{r}:{last_helper_letter}{r},2)>0,"ACCEPTABLE","PASS"))'
        )

    # Summary table
    start_row = 4
    start_col = 1
    ws.cell(start_row, start_col, "Summary").font = Font(bold=True, size=12)
    ws.cell(start_row+1, start_col, "Pass Count")
    ws.cell(start_row+2, start_col, "Acceptable")
    ws.cell(start_row+3, start_col, "Fail Count")
    ws.cell(start_row+4, start_col, "Total Samples")
    ws.cell(start_row+5, start_col, "Fail %")

    result_letter = get_column_letter(result_col)
    ws.cell(start_row+1, start_col+1, f'=COUNTIF({result_letter}{header_row+1}:{result_letter}{ws.max_row},"PASS")')
    ws.cell(start_row+2, start_col+1, f'=COUNTIF({result_letter}{header_row+1}:{result_letter}{ws.max_row},"ACCEPTABLE")')
    ws.cell(start_row+3, start_col+1, f'=COUNTIF({result_letter}{header_row+1}:{result_letter}{ws.max_row},"FAIL")')
    ws.cell(start_row+4, start_col+1, f'=COUNTA({result_letter}{header_row+1}:{result_letter}{ws.max_row})')
    ws.cell(start_row+5, start_col+1, f'={get_column_letter(start_col+1)}{start_row+3}/{get_column_letter(start_col+1)}{start_row+4}')
    ws.cell(start_row+5, start_col+1).number_format = "0.00%"

    # Apply border to summary
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row_offset in range(0, 6):
        for col_offset in range(2):
            cell = ws[f"{get_column_letter(start_col + col_offset)}{start_row + row_offset}"]
            cell.border = border
            if row_offset == 0:
                cell.font = Font(bold=True)

    # Auto-width for summary columns
    for col_offset in range(2):
        # WIDTH REGULATION
        set_column_width(ws, start_col + col_offset)

# ---------------------------------------------------------
# APPLY CONDITIONAL FORMATTING TO DATA TABLE
# ---------------------------------------------------------
def apply_conditional_formatting(ws, vendor):
    first_data_row = HEADER_ROW + 1
    last_data_row = ws.max_row

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 

    tol_cell = LIMIT_MAP[vendor]["TOL"]

    # Apply formatting to measurement columns
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(HEADER_ROW, c).value).upper()
        if name in ("NO", "SN", "LENS SN", "RESULT") or not name:
            continue

        col_letter = get_column_letter(c)
        limit_cell = None
        for key, cell in LIMIT_MAP[vendor].items():
            if key != "TOL" and name.startswith(key):
                limit_cell = cell
                break
        if not limit_cell:
            continue

        cell_range = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"

        # Conditional rules
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{col_letter}{first_data_row} < ({limit_cell}-{tol_cell})"], fill=red_fill, stopIfTrue=True)
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"AND({col_letter}{first_data_row} >= ({limit_cell}-{tol_cell}), {col_letter}{first_data_row} < ({limit_cell}+{tol_cell}))"],
                        fill=yellow_fill, stopIfTrue=True)
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{col_letter}{first_data_row} >= ({limit_cell}+{tol_cell})"], fill=green_fill)
        )

        # Apply border
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for r in range(HEADER_ROW, last_data_row + 1):
            cell = ws.cell(r, c)
            cell.border = border
       
# ---------------------------------------------------------
# APPLY RESULT FONTS
# ---------------------------------------------------------
def apply_result_fonts(ws):
    result_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(HEADER_ROW, c).value).upper() == "RESULT":
            result_col = c
            break
    if not result_col:
        return

    first_data_row = HEADER_ROW + 1
    last_data_row = ws.max_row
    col_letter = get_column_letter(result_col)

    # Fonts
    pass_font = Font(color="4F6228")
    fail_font = Font(color="FF0000")
    acceptable_font = Font(color="E36C09")

    # FAIL
    fail_rule = FormulaRule(
        formula=[f'{col_letter}{first_data_row}="FAIL"'],
        font=fail_font,
        stopIfTrue=True
    )
    ws.conditional_formatting.add(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}", fail_rule)

    # ACCEPTABLE
    acceptable_rule = FormulaRule(
        formula=[f'{col_letter}{first_data_row}="ACCEPTABLE"'],
        font=acceptable_font,
        stopIfTrue=True
    )
    ws.conditional_formatting.add(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}", acceptable_rule)

    # PASS
    pass_rule = FormulaRule(
        formula=[f'{col_letter}{first_data_row}="PASS"'],
        font=pass_font
    )
    ws.conditional_formatting.add(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}", pass_rule)

    # WIDTH REGULATION
    set_column_width(ws, result_col, max_width=16)

# ---------------------------------------------------------
# PROCESS FILE
# ---------------------------------------------------------
def process_file(input_file,vendor):
    df = pd.read_excel(input_file, sheet_name=0)
    if vendor=="1": out=process_vendor1(df)
    else: out=process_vendor2(df)
    
    out_file = os.path.splitext(input_file)[0]+"_processed.xlsx"
    out.to_excel(out_file, index=False, sheet_name="MTF Data")
    
    wb = load_workbook(out_file)
    ws = wb["MTF Data"]
    
    insert_limits_table(ws,vendor)
    fix_lens_sn_format(ws)
    insert_excel_formulas(ws,vendor)
    apply_conditional_formatting(ws,vendor)
    apply_result_fonts(ws)
    
    wb.save(out_file)
    print("Done:", out_file)

# ---------------------------------------------------------
# MAIN
# ---------------------------------------------------------
def main():
    print("Select vendor type:")
    print(f"1. {VENDOR1}")
    print(f"2. {VENDOR2}")
    vendor = input("Enter 1 or 2: ").strip()
    if vendor not in ("1","2"): sys.exit("Invalid vendor")
    input_file = input("Path to .xlsx file: ").strip('"').strip("'")
    if not os.path.exists(input_file): sys.exit("File not found")
    process_file(input_file,vendor)

if __name__=="__main__":
    main()
