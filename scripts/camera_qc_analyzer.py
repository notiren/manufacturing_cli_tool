#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║         CAMERA QC ANALYZER v4.0 — Desktop Tool              ║
║         BlackNoise + IR Cut Analysis Pipeline                ║
║                                                              ║
║  Workflow:                                                   ║
║    1. Select root folder containing subfolders:              ║
║         BlackNoisePicUrl/                                    ║
║         IrCutOnFirstPicUrl/                                  ║
║         IrCutOnSecondPicUrl/                                 ║
║         IrCutOffFirstPicUrl/                                 ║
║         IrCutOffSecondPicUrl/                                ║
║    2. Set thresholds for BlackNoise & IR Cut via UI          ║
║    3. Run analysis → view results per test                   ║
║    4. Export Excel report + Split PASS/FAIL folders          ║
╚══════════════════════════════════════════════════════════════╝


Metrics:
    BlackNoise:  Max-Channel Mean (V) — mean of max(R,G,B) per pixel
                 brightness < threshold → PASS
    IR Cut:      R-G Difference — mean(R) - mean(G) per image
                 IrCutOff:  R-G diff should be NEGATIVE (normal) → R-G < threshold → PASS
                 IrCutOn:   R-G diff should be POSITIVE (pinkish) → R-G >= threshold → PASS
"""

import os
import sys
import time
import shutil
import threading
import platform
import subprocess
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── Dependency check ────────────────────────────────────────────────────────
def ensure_package(pkg, imp=None):
    try:
        __import__(imp or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", pkg])

ensure_package("cv2")
ensure_package("pandas")
ensure_package("numpy")
ensure_package("openpyxl")

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ── Constants ───────────────────────────────────────────────────────────────
SUPPORTED_EXT = {'.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif', '.webp'}

# Expected subfolder names
FOLDER_BLACKNOISE     = "BlackNoisePicUrl"
FOLDER_IRCUT_ON_1ST   = "IrCutOnFirstPicUrl"
FOLDER_IRCUT_ON_2ND   = "IrCutOnSecondPicUrl"
FOLDER_IRCUT_OFF_1ST  = "IrCutOffFirstPicUrl"
FOLDER_IRCUT_OFF_2ND  = "IrCutOffSecondPicUrl"

ALL_SUBFOLDERS = [
    FOLDER_BLACKNOISE,
    FOLDER_IRCUT_ON_1ST, FOLDER_IRCUT_ON_2ND,
    FOLDER_IRCUT_OFF_1ST, FOLDER_IRCUT_OFF_2ND,
]

# UI Colors
BG           = "#111827"
BG_CARD      = "#1f2937"
BG_TABLE     = "#0f172a"
BG_INPUT     = "#0f172a"
FG           = "#e5e7eb"
FG_DIM       = "#9ca3af"
ACCENT       = "#3b82f6"
ACCENT_HOVER = "#2563eb"
PASS_FG      = "#22c55e"
FAIL_FG      = "#ef4444"
WARN_FG      = "#f59e0b"
BORDER_CLR   = "#374151"
SPLIT_CLR    = "#f59e0b"
SPLIT_HOVER  = "#d97706"
TAB_ACTIVE   = "#3b82f6"
TAB_INACTIVE = "#374151"
TERM_BG      = "#0c0c0c"
TERM_FG      = "#cccccc"
TERM_GREEN   = "#16c60c"
TERM_RED     = "#e74856"
TERM_YELLOW  = "#f9f1a5"
TERM_BLUE    = "#3b78ff"
TERM_CYAN    = "#61d6d6"
TERM_MAGENTA = "#b4009e"
TERM_DIM     = "#767676"
TERM_WHITE   = "#f2f2f2"


# ── Helpers ─────────────────────────────────────────────────────────────────
def _imread(filepath):
    """cv2.imread wrapper that handles non-ASCII/Unicode paths on Windows."""
    buf = np.fromfile(filepath, dtype=np.uint8)
    return cv2.imdecode(buf, cv2.IMREAD_COLOR)


# ── Analysis Functions ──────────────────────────────────────────────────────
def analyze_blacknoise(filepath):
    """Brightness analysis for black-noise images."""
    img = _imread(filepath)
    if img is None:
        return None
    # Compute V = max(R,G,B) per pixel using numpy vectorized ops on raw array
    # img shape is (H,W,3) in BGR order
    v_channel = img.max(axis=2)
    brightness = float(v_channel.mean())
    # Grayscale via weighted sum (matches cv2 BT.601) — avoids cvtColor call
    # gray = 0.114*B + 0.587*G + 0.299*R
    b_mean, g_mean, r_mean, _ = cv2.mean(img)
    grayscale_mean = 0.299 * r_mean + 0.587 * g_mean + 0.114 * b_mean
    std_dev = float(v_channel.astype(np.float32).std())
    h, w = img.shape[:2]
    return {
        'brightness':     round(brightness, 2),
        'grayscale_mean': round(grayscale_mean, 2),
        'std_dev':        round(std_dev, 2),
        'v_min':          int(v_channel.min()),
        'v_max':          int(v_channel.max()),
        'resolution':     f"{w}x{h}",
    }


def analyze_ircut(filepath):
    """Color-cast analysis for IR cut images."""
    img = _imread(filepath)
    if img is None:
        return None
    # Use cv2.mean for fast channel means (C++ optimized, no python array alloc)
    b_mean, g_mean, r_mean, _ = cv2.mean(img)
    rg_diff = r_mean - g_mean
    magenta_dev = (r_mean + b_mean) / 2.0 - g_mean

    # Pink pixel % — detect from BGR directly, avoid full HSV conversion
    # Pink in BGR: R high, G low relative to R, B moderate
    b, g, r = img[:,:,0], img[:,:,1], img[:,:,2]
    r_i = r.astype(np.int16)
    g_i = g.astype(np.int16)
    pink_mask = (r > 80) & ((r_i - g_i) > 20) & (r > b)
    pink_pct = 100.0 * float(pink_mask.sum()) / pink_mask.size

    height, width = img.shape[:2]
    return {
        'r_mean':      round(r_mean, 2),
        'g_mean':      round(g_mean, 2),
        'b_mean':      round(b_mean, 2),
        'rg_diff':     round(rg_diff, 2),
        'magenta_dev': round(magenta_dev, 2),
        'pink_pct':    round(pink_pct, 2),
        'resolution':  f"{width}x{height}",
    }


# ── Excel Export ────────────────────────────────────────────────────────────
def export_full_report(all_results, thresholds, output_path):
    """
    Export a multi-sheet Excel report.
    all_results: dict with keys like 'BlackNoisePicUrl' -> list of result dicts
    thresholds: dict with 'blacknoise' and 'ircut' values
    """
    wb = Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1e293b")
    data_font = Font(name="Arial", size=10)
    pass_fill = PatternFill("solid", fgColor="dcfce7")
    fail_fill = PatternFill("solid", fgColor="fecaca")
    pass_font = Font(name="Arial", size=10, bold=True, color="166534")
    fail_font = Font(name="Arial", size=10, bold=True, color="991b1b")
    border = Border(
        left=Side(style='thin', color='d1d5db'),
        right=Side(style='thin', color='d1d5db'),
        top=Side(style='thin', color='d1d5db'),
        bottom=Side(style='thin', color='d1d5db'),
    )
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    first_sheet = True

    # ── Per-device summary sheet (overall) ──
    ws_summary = wb.active
    ws_summary.title = "Device Summary"

    # Gather all device IDs across all folders
    device_map = {}  # device_id -> {folder_name: status}
    for folder_name, results in all_results.items():
        for r in results:
            dev_id = r['filename'].replace('.png', '').replace('.jpg', '').replace('.jpeg', '')
            if dev_id not in device_map:
                device_map[dev_id] = {}
            device_map[dev_id][folder_name] = r['status']

    # Build summary headers
    present_folders = [f for f in ALL_SUBFOLDERS if f in all_results and all_results[f]]
    summary_headers = ["SN", "Device ID"] + present_folders + ["Overall"]
    summary_widths = [8, 20] + [22] * len(present_folders) + [12]

    for ci, (h, w) in enumerate(zip(summary_headers, summary_widths), 1):
        cell = ws_summary.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws_summary.column_dimensions[get_column_letter(ci)].width = w

    for ri, (dev_id, statuses) in enumerate(sorted(device_map.items()), 2):
        overall = "PASS" if all(statuses.get(f) == "PASS" for f in present_folders) else "FAIL"
        vals = [ri - 1, dev_id] + [statuses.get(f, "N/A") for f in present_folders] + [overall]
        alt_fill = PatternFill("solid", fgColor="f8fafc") if ri % 2 == 0 else None
        for ci, v in enumerate(vals, 1):
            cell = ws_summary.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.alignment = center
            cell.border = border
            if v == "PASS":
                cell.fill = pass_fill
                cell.font = pass_font
            elif v == "FAIL":
                cell.fill = fail_fill
                cell.font = fail_font
            elif v == "N/A":
                cell.font = Font(name="Arial", size=10, color="9ca3af")
            elif alt_fill:
                cell.fill = alt_fill

    ws_summary.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}{len(device_map) + 1}"
    ws_summary.freeze_panes = "A2"

    # ── Detail sheet per folder ──
    for folder_name in ALL_SUBFOLDERS:
        if folder_name not in all_results or not all_results[folder_name]:
            continue

        results = all_results[folder_name]
        is_blacknoise = ("BlackNoise" in folder_name)

        ws = wb.create_sheet(title=folder_name[:31])  # sheet name max 31 chars

        if is_blacknoise:
            headers = ["SN", "Filename", "Brightness (V)", "Grayscale", "Std Dev",
                       "V Min", "V Max", "Resolution", "Status"]
            widths = [8, 22, 16, 14, 12, 10, 10, 14, 10]
        else:
            headers = ["SN", "Filename", "R Mean", "G Mean", "B Mean", "R-G Diff",
                       "Magenta Dev", "Pink %", "Resolution", "Status"]
            widths = [8, 22, 12, 12, 12, 12, 14, 10, 14, 10]

        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, item in enumerate(results, 2):
            if is_blacknoise:
                vals = [item['sn'], item['filename'], item['brightness'],
                        item['grayscale_mean'], item['std_dev'], item['v_min'],
                        item['v_max'], item['resolution'], item['status']]
            else:
                vals = [item['sn'], item['filename'], item['r_mean'], item['g_mean'],
                        item['b_mean'], item['rg_diff'], item['magenta_dev'],
                        item['pink_pct'], item['resolution'], item['status']]

            alt_fill = PatternFill("solid", fgColor="f8fafc") if ri % 2 == 0 else None
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
                if ci == len(vals):
                    cell.fill = pass_fill if v == "PASS" else fail_fill
                    cell.font = pass_font if v == "PASS" else fail_font
                elif alt_fill:
                    cell.fill = alt_fill

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"
        ws.freeze_panes = "A2"

    # ── Thresholds & stats sheet ──
    ws_info = wb.create_sheet("Settings & Stats")
    ws_info.column_dimensions['A'].width = 28
    ws_info.column_dimensions['B'].width = 20

    info_rows = [
        ("Setting", "Value"),
        ("BlackNoise Threshold", thresholds.get('blacknoise', 'N/A')),
        ("IR Cut Threshold (R-G diff)", thresholds.get('ircut', 'N/A')),
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
    ]

    for folder_name in ALL_SUBFOLDERS:
        if folder_name not in all_results or not all_results[folder_name]:
            continue
        results = all_results[folder_name]
        passed = sum(1 for r in results if r['status'] == 'PASS')
        failed = len(results) - passed
        info_rows.append((f"--- {folder_name} ---", ""))
        info_rows.append(("  Total Images", len(results)))
        info_rows.append(("  PASS", passed))
        info_rows.append(("  FAIL", failed))
        info_rows.append(("  Pass Rate %", round(100 * passed / max(len(results), 1), 1)))
        info_rows.append(("", ""))

    for ri, (label, val) in enumerate(info_rows, 1):
        ca = ws_info.cell(row=ri, column=1, value=label)
        cb = ws_info.cell(row=ri, column=2, value=val)
        if ri == 1:
            ca.font = hdr_font
            ca.fill = hdr_fill
            cb.font = hdr_font
            cb.fill = hdr_fill
        else:
            ca.font = Font(name="Arial", size=10, bold=True)
            cb.font = data_font
        ca.alignment = center
        cb.alignment = center
        ca.border = border
        cb.border = border

    wb.save(output_path)


# ── Folder Split ────────────────────────────────────────────────────────────
def split_results_to_folders(all_results, root_folder, output_base):
    """
    Create PASS/ and FAIL/ folders under output_base.
    Inside each, replicate the subfolder structure.
    """
    pass_root = os.path.join(output_base, "PASS")
    fail_root = os.path.join(output_base, "FAIL")
    copied = 0
    errors = []

    for folder_name, results in all_results.items():
        if not results:
            continue
        pass_dir = os.path.join(pass_root, folder_name)
        fail_dir = os.path.join(fail_root, folder_name)
        os.makedirs(pass_dir, exist_ok=True)
        os.makedirs(fail_dir, exist_ok=True)

        src_dir = os.path.join(root_folder, folder_name)
        for r in results:
            src = os.path.join(src_dir, r['filename'])
            dst_dir = pass_dir if r['status'] == "PASS" else fail_dir
            dst = os.path.join(dst_dir, r['filename'])
            try:
                shutil.copy2(src, dst)
                copied += 1
            except Exception as e:
                errors.append(f"{folder_name}/{r['filename']}: {e}")

    return pass_root, fail_root, copied, errors


# ══════════════════════════════════════════════════════════════════════════════
# ── Threshold Configuration Dialog ───────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
class ThresholdConfigDialog(tk.Toplevel):
    """
    Modal dialog shown before analysis.
    Lets user set BlackNoise threshold AND IR Cut threshold.
    Shows which subfolders were detected and image counts.
    """

    def __init__(self, parent, bn_threshold, ircut_threshold, detected_folders, root_folder):
        super().__init__(parent)
        self.result = None

        self.title("Configure Analysis Thresholds")
        self.configure(bg=BG)
        self.resizable(False, False)

        w, h = 560, 560
        px = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        py = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{px}+{py}")

        self.transient(parent)
        self.grab_set()

        self._bn_var = tk.DoubleVar(value=bn_threshold)
        self._ircut_var = tk.DoubleVar(value=ircut_threshold)

        # ── Header ──
        tk.Label(self, text="⚙  Analysis Configuration",
                 font=("Segoe UI", 16, "bold"),
                 bg=BG, fg=ACCENT).pack(pady=(20, 4))

        folder_name = os.path.basename(root_folder)
        tk.Label(self, text=f"Root: {folder_name}",
                 font=("Segoe UI", 9), bg=BG, fg=FG_DIM).pack(pady=(0, 12))

        # ── Detected folders ──
        det_frame = tk.Frame(self, bg=BG_CARD, highlightbackground=BORDER_CLR, highlightthickness=1)
        det_frame.pack(padx=30, fill="x", ipady=8)

        tk.Label(det_frame, text="DETECTED SUBFOLDERS",
                 font=("Segoe UI", 9, "bold"), bg=BG_CARD, fg=FG_DIM).pack(pady=(8, 4))

        for fname in ALL_SUBFOLDERS:
            count = detected_folders.get(fname, 0)
            color = PASS_FG if count > 0 else FAIL_FG
            icon = "✓" if count > 0 else "✗"
            text = f"  {icon}  {fname}  —  {count} images" if count > 0 else f"  {icon}  {fname}  —  not found"
            tk.Label(det_frame, text=text, font=("Consolas", 9), bg=BG_CARD, fg=color,
                     anchor="w").pack(fill="x", padx=16)

        tk.Label(det_frame, text="", bg=BG_CARD).pack(pady=(0, 4))

        # ── BlackNoise threshold ──
        bn_card = tk.Frame(self, bg=BG_CARD, highlightbackground=BORDER_CLR, highlightthickness=1)
        bn_card.pack(padx=30, pady=(12, 0), fill="x", ipady=6)

        bn_row = tk.Frame(bn_card, bg=BG_CARD)
        bn_row.pack(fill="x", padx=16, pady=8)
        tk.Label(bn_row, text="BLACKNOISE THRESHOLD",
                 font=("Segoe UI", 10, "bold"), bg=BG_CARD, fg=FG).pack(side="left")

        self._bn_entry = tk.Entry(bn_row, textvariable=self._bn_var, width=8,
                                  font=("Consolas", 14, "bold"), justify="center",
                                  bg=BG_INPUT, fg=ACCENT, insertbackground=ACCENT,
                                  relief="flat", borderwidth=2)
        self._bn_entry.pack(side="right", ipady=3)

        tk.Label(bn_card, text="V-mean < threshold → PASS  (lower = darker = better)",
                 font=("Segoe UI", 8), bg=BG_CARD, fg=FG_DIM).pack(padx=16, anchor="w", pady=(0, 6))

        # ── IR Cut threshold ──
        ir_card = tk.Frame(self, bg=BG_CARD, highlightbackground=BORDER_CLR, highlightthickness=1)
        ir_card.pack(padx=30, pady=(10, 0), fill="x", ipady=6)

        ir_row = tk.Frame(ir_card, bg=BG_CARD)
        ir_row.pack(fill="x", padx=16, pady=8)
        tk.Label(ir_row, text="IR CUT THRESHOLD  (R-G diff)",
                 font=("Segoe UI", 10, "bold"), bg=BG_CARD, fg=FG).pack(side="left")

        self._ircut_entry = tk.Entry(ir_row, textvariable=self._ircut_var, width=8,
                                     font=("Consolas", 14, "bold"), justify="center",
                                     bg=BG_INPUT, fg=WARN_FG, insertbackground=WARN_FG,
                                     relief="flat", borderwidth=2)
        self._ircut_entry.pack(side="right", ipady=3)

        tk.Label(ir_card, text="IrCutOff: R-G < threshold → PASS  (normal, not pink)\n"
                               "IrCutOn:  R-G ≥ threshold → PASS  (correctly pinkish)",
                 font=("Segoe UI", 8), bg=BG_CARD, fg=FG_DIM, justify="left").pack(padx=16, anchor="w", pady=(0, 6))

        # ── Buttons ──
        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.pack(pady=(18, 16))

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=self._cancel,
                               bg=BG_CARD, fg=FG_DIM, activebackground=BORDER_CLR,
                               activeforeground="#ffffff",
                               font=("Segoe UI", 10, "bold"), relief="flat",
                               cursor="hand2", padx=24, pady=8, borderwidth=0)
        cancel_btn.pack(side="left", padx=(0, 12))

        run_btn = tk.Button(btn_frame, text="▶  Start Full Analysis", command=self._confirm,
                            bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HOVER,
                            activeforeground="#ffffff",
                            font=("Segoe UI", 11, "bold"), relief="flat",
                            cursor="hand2", padx=28, pady=8, borderwidth=0)
        run_btn.pack(side="left")

        self.bind("<Return>", lambda e: self._confirm())
        self.bind("<Escape>", lambda e: self._cancel())
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self._bn_entry.focus_set()
        self._bn_entry.select_range(0, tk.END)
        self.wait_window()

    def _confirm(self):
        try:
            bn = self._bn_var.get()
            ir = self._ircut_var.get()
            if bn <= 0 or bn > 255:
                messagebox.showwarning("Invalid", "BlackNoise threshold must be 0.1–255.", parent=self)
                return
            self.result = {'blacknoise': bn, 'ircut': ir}
        except (tk.TclError, ValueError):
            messagebox.showwarning("Invalid", "Please enter valid numbers.", parent=self)
            return
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# ── Main GUI ─────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
class CameraQCApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Camera QC Analyzer v4.0 — BlackNoise + IR Cut")
        self.root.configure(bg=BG)
        self.root.minsize(1080, 750)
        try:
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            w, h = min(1200, sw - 80), min(820, sh - 80)
            x, y = (sw - w) // 2, (sh - h) // 2
            self.root.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            self.root.geometry("1200x820")

        self.root_folder = tk.StringVar(value="")
        self.bn_threshold = tk.DoubleVar(value=45.0)
        self.ircut_threshold = tk.DoubleVar(value=-4.0)
        self.all_results = {}     # folder_name -> [result dicts]
        self.detected_folders = {}  # folder_name -> image count
        self.running = False

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Main.TFrame",       background=BG)
        style.configure("Card.TFrame",       background=BG_CARD)
        style.configure("Title.TLabel",      background=BG, foreground=ACCENT, font=("Segoe UI", 18, "bold"))
        style.configure("Subtitle.TLabel",   background=BG, foreground=FG_DIM, font=("Segoe UI", 10))
        style.configure("CardTitle.TLabel",  background=BG_CARD, foreground=FG, font=("Segoe UI", 11, "bold"))
        style.configure("Info.TLabel",       background=BG_CARD, foreground=FG_DIM, font=("Segoe UI", 9))
        style.configure("Path.TLabel",       background=BG_CARD, foreground=ACCENT, font=("Segoe UI", 9))
        style.configure("Status.TLabel",     background=BG, foreground=FG_DIM, font=("Segoe UI", 9))
        style.configure("StatVal.TLabel",    background=BG_CARD, foreground=FG, font=("Segoe UI", 16, "bold"))
        style.configure("StatLabel.TLabel",  background=BG_CARD, foreground=FG_DIM, font=("Segoe UI", 8))
        style.configure("PassVal.TLabel",    background=BG_CARD, foreground=PASS_FG, font=("Segoe UI", 16, "bold"))
        style.configure("FailVal.TLabel",    background=BG_CARD, foreground=FAIL_FG, font=("Segoe UI", 16, "bold"))
        style.configure("Accent.Horizontal.TProgressbar",
                         troughcolor=BG_CARD, background=ACCENT,
                         darkcolor=ACCENT, lightcolor=ACCENT, bordercolor=BG_CARD)
        style.configure("Custom.Treeview",
                         background=BG_TABLE, foreground=FG, fieldbackground=BG_TABLE,
                         rowheight=24, font=("Consolas", 9), borderwidth=0)
        style.configure("Custom.Treeview.Heading",
                         background="#1e293b", foreground=ACCENT,
                         font=("Segoe UI", 9, "bold"), borderwidth=0)
        style.map("Custom.Treeview",
                   background=[("selected", ACCENT)], foreground=[("selected", "#ffffff")])

    def _make_button(self, parent, text, command, bg_clr=None, hover_clr=None, fg_clr=None):
        bg = bg_clr or ACCENT
        fg = fg_clr or "#ffffff"
        hover = hover_clr or ACCENT_HOVER
        btn = tk.Button(parent, text=text, command=command,
                        bg=bg, fg=fg, activebackground=hover, activeforeground="#ffffff",
                        font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                        padx=14, pady=7, borderwidth=0)
        btn.bind("<Enter>", lambda e: btn.configure(bg=hover))
        btn.bind("<Leave>", lambda e: btn.configure(bg=bg))
        return btn

    def _build_ui(self):
        # ── Header ──
        hdr = ttk.Frame(self.root, style="Main.TFrame")
        hdr.pack(fill="x", padx=20, pady=(14, 2))
        ttk.Label(hdr, text="⬡  Camera QC Analyzer v4", style="Title.TLabel").pack(side="left")
        ttk.Label(hdr, text="BlackNoise + IR Cut Pipeline",
                  style="Subtitle.TLabel").pack(side="left", padx=(12, 0), pady=(5, 0))

        # ── Control bar ──
        ctrl = ttk.Frame(self.root, style="Main.TFrame")
        ctrl.pack(fill="x", padx=20, pady=(8, 0))

        folder_card = ttk.Frame(ctrl, style="Card.TFrame")
        folder_card.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)
        top_row = ttk.Frame(folder_card, style="Card.TFrame")
        top_row.pack(fill="x", padx=10, pady=(6, 2))
        ttk.Label(top_row, text="ROOT FOLDER", style="CardTitle.TLabel").pack(side="left")
        self.browse_btn = self._make_button(top_row, "📁  Browse", self._browse_folder)
        self.browse_btn.pack(side="right")
        self.folder_label = ttk.Label(folder_card, text="No folder selected", style="Path.TLabel")
        self.folder_label.pack(fill="x", padx=10, pady=(0, 2))
        self.detect_label = ttk.Label(folder_card, text="", style="Info.TLabel")
        self.detect_label.pack(fill="x", padx=10)

        # Threshold display & run
        right_card = ttk.Frame(ctrl, style="Card.TFrame")
        right_card.pack(side="right", ipadx=10, ipady=8, padx=(10, 0))
        ttk.Label(right_card, text="THRESHOLDS", style="CardTitle.TLabel").pack(padx=10, pady=(6, 4))

        thresh_info = ttk.Frame(right_card, style="Card.TFrame")
        thresh_info.pack(padx=10)
        self.bn_display = tk.Label(thresh_info, text=f"BlackNoise: {self.bn_threshold.get()}",
                                   font=("Consolas", 10, "bold"), bg=BG_CARD, fg=ACCENT)
        self.bn_display.pack(anchor="w")
        self.ir_display = tk.Label(thresh_info, text=f"IR Cut R-G: {self.ircut_threshold.get()}",
                                   font=("Consolas", 10, "bold"), bg=BG_CARD, fg=WARN_FG)
        self.ir_display.pack(anchor="w")

        self.run_btn = self._make_button(right_card, "▶  Run Analysis", self._start_analysis)
        self.run_btn.pack(padx=10, pady=(8, 4))

        # ── Progress ──
        prog_frame = ttk.Frame(self.root, style="Main.TFrame")
        prog_frame.pack(fill="x", padx=20, pady=(8, 0))
        self.progress_bar = ttk.Progressbar(prog_frame, mode="determinate",
                                             style="Accent.Horizontal.TProgressbar")
        self.progress_bar.pack(fill="x")
        self.progress_label = ttk.Label(prog_frame, text="", style="Status.TLabel")
        self.progress_label.pack(fill="x", pady=(3, 0))

        # ── Stats bar ──
        self.stats_frame = ttk.Frame(self.root, style="Main.TFrame")

        # ── Tab bar ──
        self.tab_bar = tk.Frame(self.root, bg=BG)
        self.tab_bar.pack(fill="x", padx=20, pady=(8, 0))
        self.tab_buttons = {}

        # ── Terminal log (main content area) ──
        self._build_terminal()

        # ── Bottom bar ──
        bottom = ttk.Frame(self.root, style="Main.TFrame")
        bottom.pack(fill="x", padx=20, pady=(6, 12))

        self.export_btn = self._make_button(bottom, "💾  Export Excel Report", self._export_excel)
        self.export_btn.pack(side="right")
        self.export_btn.configure(state="disabled")

        self.split_btn = self._make_button(bottom, "📂  Split PASS / FAIL Folders", self._split_folders,
                                           bg_clr=SPLIT_CLR, hover_clr=SPLIT_HOVER)
        self.split_btn.pack(side="right", padx=(0, 8))
        self.split_btn.configure(state="disabled")

        self.rerun_btn = self._make_button(bottom, "🔄  Re-classify", self._rerun,
                                           bg_clr=BG_CARD, hover_clr=BORDER_CLR, fg_clr=FG_DIM)
        self.rerun_btn.pack(side="right", padx=(0, 8))
        self.rerun_btn.configure(state="disabled")

        self.bottom_status = ttk.Label(bottom, text="Select a root folder to begin", style="Status.TLabel")
        self.bottom_status.pack(side="left")

    def _build_terminal(self):
        """Build the terminal-style log panel."""
        term_frame = tk.Frame(self.root, bg=TERM_BG)
        term_frame.pack(fill="both", expand=True, padx=20, pady=(4, 0))

        # Terminal header bar
        term_hdr = tk.Frame(term_frame, bg="#1a1a1a")
        term_hdr.pack(fill="x")

        tk.Label(term_hdr, text="⬤", font=("Segoe UI", 7), bg="#1a1a1a",
                 fg=TERM_RED).pack(side="left", padx=(8, 0), pady=3)
        tk.Label(term_hdr, text="⬤", font=("Segoe UI", 7), bg="#1a1a1a",
                 fg=TERM_YELLOW).pack(side="left", padx=(3, 0), pady=3)
        tk.Label(term_hdr, text="⬤", font=("Segoe UI", 7), bg="#1a1a1a",
                 fg=TERM_GREEN).pack(side="left", padx=(3, 0), pady=3)
        tk.Label(term_hdr, text="  TERMINAL — Analysis Log",
                 font=("Consolas", 9, "bold"), bg="#1a1a1a",
                 fg=TERM_DIM).pack(side="left", padx=(6, 0), pady=3)

        clear_btn = tk.Button(term_hdr, text="Clear", font=("Consolas", 8),
                              bg="#1a1a1a", fg=TERM_DIM, activebackground="#333",
                              activeforeground=TERM_WHITE, relief="flat", cursor="hand2",
                              borderwidth=0, command=self._clear_terminal)
        clear_btn.pack(side="right", padx=8, pady=3)

        # Terminal text widget
        term_body = tk.Frame(term_frame, bg=TERM_BG)
        term_body.pack(fill="both", expand=True)

        self.terminal = tk.Text(term_body, bg=TERM_BG, fg=TERM_FG,
                                font=("Consolas", 9), relief="flat",
                                insertbackground=TERM_GREEN, selectbackground="#264f78",
                                selectforeground=TERM_WHITE, borderwidth=0,
                                wrap="word", state="disabled", padx=10, pady=6)
        term_scroll = ttk.Scrollbar(term_body, orient="vertical", command=self.terminal.yview)
        self.terminal.configure(yscrollcommand=term_scroll.set)
        self.terminal.pack(side="left", fill="both", expand=True)
        term_scroll.pack(side="right", fill="y")

        # Configure color tags
        self.terminal.tag_configure("info",    foreground=TERM_BLUE)
        self.terminal.tag_configure("success", foreground=TERM_GREEN)
        self.terminal.tag_configure("error",   foreground=TERM_RED)
        self.terminal.tag_configure("warn",    foreground=TERM_YELLOW)
        self.terminal.tag_configure("dim",     foreground=TERM_DIM)
        self.terminal.tag_configure("cyan",    foreground=TERM_CYAN)
        self.terminal.tag_configure("magenta", foreground=TERM_MAGENTA)
        self.terminal.tag_configure("white",   foreground=TERM_WHITE)
        self.terminal.tag_configure("pass_tag", foreground=TERM_GREEN, font=("Consolas", 9, "bold"))
        self.terminal.tag_configure("fail_tag", foreground=TERM_RED, font=("Consolas", 9, "bold"))
        self.terminal.tag_configure("header",  foreground=TERM_CYAN, font=("Consolas", 9, "bold"))
        self.terminal.tag_configure("timestamp", foreground=TERM_DIM)

    def _log(self, message, tag="white", timestamp=True):
        """Append a message to the terminal log with optional color tag and timestamp."""
        self.terminal.configure(state="normal")
        if timestamp:
            ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            self.terminal.insert("end", f"[{ts}] ", "timestamp")
        self.terminal.insert("end", message + "\n", tag)
        self.terminal.see("end")
        self.terminal.configure(state="disabled")

    def _log_safe(self, message, tag="white", timestamp=True):
        """Thread-safe log: schedule _log on the main thread."""
        self.root.after(0, self._log, message, tag, timestamp)

    def _clear_terminal(self):
        self.terminal.configure(state="normal")
        self.terminal.delete("1.0", "end")
        self.terminal.configure(state="disabled")

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Select root folder with image subfolders")
        if not folder:
            return
        self.root_folder.set(folder)
        self.folder_label.configure(text=folder)
        self._log(f"Root folder set → {folder}", "info")

        # Detect subfolders
        self.detected_folders = {}
        found = 0
        self._log("Scanning for subfolders...", "dim")
        for sf in ALL_SUBFOLDERS:
            sf_path = os.path.join(folder, sf)
            if os.path.isdir(sf_path):
                imgs = [f for f in os.listdir(sf_path)
                        if os.path.isfile(os.path.join(sf_path, f)) and
                        os.path.splitext(f)[1].lower() in SUPPORTED_EXT]
                self.detected_folders[sf] = len(imgs)
                found += len(imgs)
                self._log(f"  ✓ {sf}: {len(imgs)} images", "success")
            else:
                self.detected_folders[sf] = 0
                self._log(f"  ✗ {sf}: not found", "dim")

        detected_names = [k for k, v in self.detected_folders.items() if v > 0]
        if detected_names:
            self.detect_label.configure(
                text=f"Found {len(detected_names)} subfolder(s), {found} total images"
            )
            self._log(f"Ready — {found} images across {len(detected_names)} folders", "cyan")
        else:
            self.detect_label.configure(text="⚠  No recognized subfolders found")
            self._log("WARNING: No recognized subfolders found!", "warn")

        self.bottom_status.configure(text=f"Ready — {found} images across {len(detected_names)} folders")

    def _start_analysis(self):
        folder = self.root_folder.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("No Folder", "Please select a root folder first.")
            return
        if not any(v > 0 for v in self.detected_folders.values()):
            messagebox.showwarning("No Images", "No recognized subfolders with images found.")
            return

        # ── Show threshold config dialog ──
        dialog = ThresholdConfigDialog(
            self.root, self.bn_threshold.get(), self.ircut_threshold.get(),
            self.detected_folders, folder
        )
        if dialog.result is None:
            return

        self.bn_threshold.set(dialog.result['blacknoise'])
        self.ircut_threshold.set(dialog.result['ircut'])
        self.bn_display.configure(text=f"BlackNoise: {dialog.result['blacknoise']}")
        self.ir_display.configure(text=f"IR Cut R-G: {dialog.result['ircut']}")

        # Start threaded analysis
        self.running = True
        self.run_btn.configure(state="disabled")
        self.browse_btn.configure(state="disabled")
        self.export_btn.configure(state="disabled")
        self.split_btn.configure(state="disabled")
        self.rerun_btn.configure(state="disabled")
        self.all_results = {}
        self.stats_frame.pack_forget()

        # Count total images
        total = sum(v for v in self.detected_folders.values() if v > 0)
        self.progress_bar["maximum"] = total
        self.progress_bar["value"] = 0

        self._log("═" * 60, "dim", timestamp=False)
        self._log("ANALYSIS STARTED", "header")
        self._log(f"  BlackNoise threshold : {dialog.result['blacknoise']}", "info")
        self._log(f"  IR Cut threshold     : {dialog.result['ircut']}", "warn")
        self._log(f"  Total images         : {total}", "white")
        self._log("═" * 60, "dim", timestamp=False)

        thread = threading.Thread(target=self._run_full_analysis, daemon=True)
        thread.start()

    def _run_full_analysis(self):
        folder = self.root_folder.get()
        bn_thresh = self.bn_threshold.get()
        ir_thresh = self.ircut_threshold.get()

        global_idx = 0
        total = sum(v for v in self.detected_folders.values() if v > 0)
        start = time.time()
        num_workers = min(os.cpu_count() or 4, 8)
        self._log_safe(f"Using {num_workers} worker threads", "dim")

        for subfolder in ALL_SUBFOLDERS:
            if self.detected_folders.get(subfolder, 0) == 0:
                continue

            sf_path = os.path.join(folder, subfolder)
            files = sorted([f for f in os.listdir(sf_path)
                            if os.path.isfile(os.path.join(sf_path, f)) and
                            os.path.splitext(f)[1].lower() in SUPPORTED_EXT])

            is_blacknoise = ("BlackNoise" in subfolder)
            is_ircut_on = ("IrCutOn" in subfolder)

            self._log_safe(f"── Processing: {subfolder} ({len(files)} files) ──", "header")

            # Analyze images in parallel using thread pool
            analyze_fn = analyze_blacknoise if is_blacknoise else analyze_ircut
            futures = {}
            with ThreadPoolExecutor(max_workers=num_workers) as executor:
                for i, fname in enumerate(files, 1):
                    fpath = os.path.join(sf_path, fname)
                    future = executor.submit(analyze_fn, fpath)
                    futures[future] = (i, fname)

                results = []
                sf_pass = 0
                sf_fail = 0
                log_batch = []
                batch_size = max(1, min(50, len(files) // 20))

                for future in as_completed(futures):
                    sn, fname = futures[future]
                    global_idx += 1
                    metrics = future.result()

                    if metrics is None:
                        log_batch.append((f"  SKIP  {fname} — could not read image", "warn"))
                        continue

                    if is_blacknoise:
                        status = "PASS" if metrics['brightness'] < bn_thresh else "FAIL"
                        val_str = f"V={metrics['brightness']:.1f}"
                    elif is_ircut_on:
                        status = "PASS" if metrics['rg_diff'] >= ir_thresh else "FAIL"
                        val_str = f"R-G={metrics['rg_diff']:.1f}"
                    else:
                        status = "PASS" if metrics['rg_diff'] < ir_thresh else "FAIL"
                        val_str = f"R-G={metrics['rg_diff']:.1f}"

                    metrics['status'] = status
                    metrics['sn'] = sn
                    metrics['filename'] = fname
                    metrics['subfolder'] = subfolder
                    results.append(metrics)

                    if status == "PASS":
                        sf_pass += 1
                        tag = "pass_tag"
                    else:
                        sf_fail += 1
                        tag = "fail_tag"

                    log_batch.append((f"  {status:4s}  {fname}  {val_str}", tag))

                    # Flush log + progress in batches to avoid GUI overhead
                    if len(log_batch) >= batch_size:
                        batch_copy = log_batch[:]
                        log_batch.clear()
                        elapsed = time.time() - start
                        speed = global_idx / elapsed if elapsed > 0 else 0
                        eta = (total - global_idx) / speed if speed > 0 else 0
                        self.root.after(0, self._flush_batch, batch_copy,
                                        global_idx, total, elapsed, eta, subfolder)

                # Flush remaining
                if log_batch:
                    elapsed = time.time() - start
                    speed = global_idx / elapsed if elapsed > 0 else 0
                    eta = (total - global_idx) / speed if speed > 0 else 0
                    self.root.after(0, self._flush_batch, log_batch,
                                    global_idx, total, elapsed, eta, subfolder)

            # Sort results by SN for consistent ordering
            results.sort(key=lambda r: r['sn'])
            self.all_results[subfolder] = results
            self._log_safe(f"  Summary: {sf_pass} PASS / {sf_fail} FAIL", "cyan")

        self.root.after(0, self._analysis_done, time.time() - start)

    def _flush_batch(self, log_entries, idx, total, elapsed, eta, subfolder):
        """Flush a batch of log entries and update progress in one GUI call."""
        self.terminal.configure(state="normal")
        for msg, tag in log_entries:
            ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            self.terminal.insert("end", f"[{ts}] ", "timestamp")
            self.terminal.insert("end", msg + "\n", tag)
        self.terminal.see("end")
        self.terminal.configure(state="disabled")

        pct = 100 * idx / total
        self.progress_bar["value"] = idx
        self.progress_label.configure(
            text=f"[{subfolder}]  {idx}/{total}  ({pct:.0f}%)  •  "
                 f"Elapsed: {elapsed:.1f}s  •  ETA: {eta:.0f}s"
        )

    def _analysis_done(self, elapsed):
        self.running = False
        self.run_btn.configure(state="normal")
        self.browse_btn.configure(state="normal")
        self.export_btn.configure(state="normal")
        self.split_btn.configure(state="normal")
        self.rerun_btn.configure(state="normal")

        total = sum(len(v) for v in self.all_results.values())
        speed = total / elapsed if elapsed > 0 else 0
        all_passed = sum(1 for v in self.all_results.values() for r in v if r['status'] == 'PASS')
        all_failed = total - all_passed

        self._log("═" * 60, "dim", timestamp=False)
        self._log("ANALYSIS COMPLETE", "header")
        self._log(f"  Total images : {total}", "white")
        self._log(f"  PASS         : {all_passed}", "success")
        self._log(f"  FAIL         : {all_failed}", "error" if all_failed > 0 else "success")
        self._log(f"  Elapsed      : {elapsed:.1f}s ({speed:.1f} img/s)", "dim")
        self._log("═" * 60, "dim", timestamp=False)

        self.progress_label.configure(
            text=f"✔  Done! {total} images in {elapsed:.1f}s ({speed:.1f} img/s)"
        )
        self.bottom_status.configure(
            text=f"Overall: {all_passed} PASS / {all_failed} FAIL  —  "
                 f"Export report or split into folders"
        )

        # Build tab bar
        self._build_tabs()
        self._show_stats_all()

    def _build_tabs(self):
        for w in self.tab_bar.winfo_children():
            w.destroy()
        self.tab_buttons = {}

        for sf in ALL_SUBFOLDERS:
            if sf not in self.all_results or not self.all_results[sf]:
                continue
            results = self.all_results[sf]
            passed = sum(1 for r in results if r['status'] == 'PASS')
            failed = len(results) - passed

            # Short label
            if "BlackNoise" in sf:
                label = f"BlackNoise ({passed}✓ {failed}✗)"
            elif "IrCutOn" in sf:
                n = "1st" if "First" in sf else "2nd"
                label = f"IR On {n} ({passed}✓ {failed}✗)"
            else:
                n = "1st" if "First" in sf else "2nd"
                label = f"IR Off {n} ({passed}✓ {failed}✗)"

            btn = tk.Button(self.tab_bar, text=label,
                            font=("Segoe UI", 9, "bold"), relief="flat",
                            cursor="hand2", padx=12, pady=5, borderwidth=0,
                            bg=TAB_INACTIVE, fg=FG)
            btn.pack(side="left", padx=(0, 3))
            self.tab_buttons[sf] = btn

    def _show_stats_all(self):
        for w in self.stats_frame.winfo_children():
            w.destroy()

        # Place stats before table
        try:
            self.stats_frame.pack(fill="x", padx=20, pady=(6, 0), before=self.tab_bar)
        except Exception:
            self.stats_frame.pack(fill="x", padx=20, pady=(6, 0))

        total = sum(len(v) for v in self.all_results.values())
        all_passed = sum(1 for v in self.all_results.values() for r in v if r['status'] == 'PASS')
        all_failed = total - all_passed

        stats = [("Total", str(total), "StatVal.TLabel")]

        for sf in ALL_SUBFOLDERS:
            if sf not in self.all_results or not self.all_results[sf]:
                continue
            res = self.all_results[sf]
            p = sum(1 for r in res if r['status'] == 'PASS')
            f = len(res) - p

            if "BlackNoise" in sf:
                short = "BN"
            elif "IrCutOn" in sf:
                short = "On" + ("1" if "First" in sf else "2")
            else:
                short = "Off" + ("1" if "First" in sf else "2")

            sty = "PassVal.TLabel" if f == 0 else "FailVal.TLabel"
            stats.append((short, f"{p}✓ {f}✗", sty))

        stats.append(("Overall", f"{all_passed}✓ {all_failed}✗",
                       "PassVal.TLabel" if all_failed == 0 else "FailVal.TLabel"))

        for label, val, sty in stats:
            card = ttk.Frame(self.stats_frame, style="Card.TFrame")
            card.pack(side="left", fill="x", expand=True, padx=(0, 6), ipady=4)
            ttk.Label(card, text=val, style=sty).pack(pady=(4, 0))
            ttk.Label(card, text=label, style="StatLabel.TLabel").pack(pady=(0, 3))

    def _rerun(self):
        """Re-classify all results with current thresholds without re-reading images."""
        if not self.all_results:
            return

        bn_thresh = self.bn_threshold.get()
        ir_thresh = self.ircut_threshold.get()

        self._log(f"Re-classifying with BN={bn_thresh}, IR={ir_thresh}...", "info")

        for sf, results in self.all_results.items():
            is_blacknoise = ("BlackNoise" in sf)
            is_ircut_on = ("IrCutOn" in sf)

            for r in results:
                if is_blacknoise:
                    r['status'] = "PASS" if r['brightness'] < bn_thresh else "FAIL"
                elif is_ircut_on:
                    r['status'] = "PASS" if r['rg_diff'] >= ir_thresh else "FAIL"
                else:
                    r['status'] = "PASS" if r['rg_diff'] < ir_thresh else "FAIL"

        self._build_tabs()
        self._show_stats_all()

        total = sum(len(v) for v in self.all_results.values())
        all_passed = sum(1 for v in self.all_results.values() for r in v if r['status'] == 'PASS')
        all_failed = total - all_passed
        self._log(f"Re-classified: {all_passed} PASS / {all_failed} FAIL", "success")
        self.bottom_status.configure(
            text=f"Re-classified: {all_passed} PASS / {all_failed} FAIL  "
                 f"(BN={bn_thresh}, IR={ir_thresh})"
        )

    def _export_excel(self):
        if not self.all_results:
            messagebox.showwarning("No Data", "Run an analysis first.")
            return
        downloads = str(Path.home() / "Downloads")
        if not os.path.isdir(downloads):
            downloads = str(Path.home())
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"camera_qc_report_{timestamp}.xlsx"
        filepath = filedialog.asksaveasfilename(
            title="Save QC Report", initialdir=downloads, initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filepath:
            return
        try:
            self._log(f"Exporting Excel report...", "info")
            thresholds = {
                'blacknoise': self.bn_threshold.get(),
                'ircut': self.ircut_threshold.get(),
            }
            export_full_report(self.all_results, thresholds, filepath)
            self._log(f"Report saved → {filepath}", "success")
            self.bottom_status.configure(text=f"Saved → {filepath}")
            if messagebox.askyesno("Export Complete", f"Report saved to:\n{filepath}\n\nOpen now?"):
                self._open_path(filepath)
        except Exception as e:
            self._log(f"Export failed: {e}", "error")
            messagebox.showerror("Export Error", f"Failed:\n{e}")

    def _split_folders(self):
        if not self.all_results:
            messagebox.showwarning("No Data", "Run an analysis first.")
            return

        folder = self.root_folder.get()
        output_dir = filedialog.askdirectory(
            title="Select where to create PASS / FAIL folders",
            initialdir=folder
        )
        if not output_dir:
            return

        total = sum(len(v) for v in self.all_results.values())
        all_passed = sum(1 for v in self.all_results.values() for r in v if r['status'] == 'PASS')
        all_failed = total - all_passed

        confirm = messagebox.askyesno(
            "Confirm Split",
            f"This will copy images into:\n\n"
            f"  PASS/  →  {all_passed} images (across subfolders)\n"
            f"  FAIL/  →  {all_failed} images (across subfolders)\n\n"
            f"Folder structure preserved inside PASS/ and FAIL/.\n"
            f"Originals will NOT be moved or deleted.\n\nContinue?"
        )
        if not confirm:
            return

        self.split_btn.configure(state="disabled")
        self.bottom_status.configure(text="Splitting images...")
        self.root.update_idletasks()
        self._log(f"Splitting {total} images into PASS/FAIL folders...", "info")

        pass_root, fail_root, copied, errors = split_results_to_folders(
            self.all_results, folder, output_dir
        )

        if errors:
            for err in errors[:5]:
                self._log(f"  ERR  {err}", "error")
            if len(errors) > 5:
                self._log(f"  ... and {len(errors) - 5} more errors", "warn")
            error_msg = "\n".join(errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... and {len(errors) - 10} more"
            messagebox.showwarning("Split Done (with errors)",
                                   f"Copied {copied}/{total}.\n\nErrors:\n{error_msg}")
        else:
            self._log(f"Split complete: {all_passed} → PASS/ • {all_failed} → FAIL/", "success")
            self.bottom_status.configure(
                text=f"✔  Split done! {all_passed} → PASS/  •  {all_failed} → FAIL/"
            )
            if messagebox.askyesno("Split Complete",
                                    f"Copied {copied} images!\n\n"
                                    f"  ✓ PASS/ → {all_passed} files\n"
                                    f"  ✗ FAIL/ → {all_failed} files\n\n"
                                    f"Open output folder?"):
                self._open_path(output_dir)

        self.split_btn.configure(state="normal")

    def _open_path(self, filepath):
        try:
            if platform.system() == "Windows":
                os.startfile(filepath)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", filepath])
            else:
                subprocess.Popen(["xdg-open", filepath])
        except Exception:
            pass


def main():
    root = tk.Tk()
    app = CameraQCApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
