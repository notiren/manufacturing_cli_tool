import sys
import subprocess

# Ensure required packages

def ensure_package(pkg, imp=None):
    try:
        __import__(imp or pkg)
    except ImportError:
        print(f"üì¶ Installing missing package: {pkg}")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", pkg])
        except subprocess.CalledProcessError as e:
            print(f"‚ùå Failed to install {pkg}: {e}")
            sys.exit(1)

ensure_package("openpyxl")
ensure_package("pandas")
ensure_package("chardet")

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import chardet

# Helper for dynamic output folder

def get_output_folder(folder_name="extracted"):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if os.path.basename(script_dir).lower() == "scripts":
        base_dir = os.path.dirname(script_dir)
    else:
        base_dir = script_dir
    output_folder = os.path.join(base_dir, folder_name)
    os.makedirs(output_folder, exist_ok=True)
    return output_folder

# CSV to Excel

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        rawdata = f.read(10000)
    result = chardet.detect(rawdata)
    encoding = result['encoding']
    return encoding if encoding else 'utf-8'

def csv_to_excel_with_headers(csv_file, output_dir):
    encoding = detect_encoding(csv_file)

    # read CSV
    df = pd.read_csv(csv_file, dtype=str, encoding=encoding)
    df.dropna(how='all', inplace=True)
    df = df.apply(lambda col: col.str.strip() if col.dtype == 'object' else col)
    df = df.loc[~(df.astype(str).apply(lambda row: row.str.strip().eq('').all(), axis=1))]

    headers = [
        "Êù°Á†Å",
        "‰∫ßÂìÅÂûãÂè∑",
        "U(V)",
        "Vp+ (Voltage peak value)",
        "Vp-\n(Voltage valley value)",
        "Vpp\n(Voltage peak valley difference)",
        "I(A)",
        "Ip+ (Peak current)",
        "Ip- (Current valley value)",
        "Ipp\n(Peak valley difference of current)",
        "P(W)",
        "Pp+ (Peak power)",
        "Pp-\n(Power valley value)",
        "Ppp\n(Rated maximum power)",
        "ÊµãËØïÊó∂Èó¥",
        "ÊµãËØï‰∫∫Âëò"
    ]

    n_cols = len(headers)
    df = df.iloc[:, :n_cols]
    df.columns = headers

    wb = Workbook()
    ws = wb.active

    # write data
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.font = Font(name='Aptos Narrow')

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_xlsx_file = os.path.join(output_dir, base_name + ".xlsx")

    try:
        wb.save(output_xlsx_file)
        print(f"-----------")
        print(f"Conversion complete. Excel file can be found inside folder: '{output_dir}'\n")
    except PermissionError:
        print(f"\nPermission denied: Could not write to '{output_dir}'. Is the file open?\n")

# ---- Main ----

def main():
    try:
        csv_file_path = input("Drop the path to a .csv file: ").strip('"').strip("'")
        
        if not csv_file_path:
            print("No file path provided. Exiting.\n")
            sys.exit(1)
        
        if not os.path.isfile(csv_file_path):
            print(f"File does not exist: {csv_file_path}\n")
            sys.exit(1)
        
        if not csv_file_path.lower().endswith('.csv'):
            print(f"The file is not a CSV: {csv_file_path}\n")
            sys.exit(1)
            
        output_path = get_output_folder("extracted")
        csv_to_excel_with_headers(csv_file_path, output_path)
        
    except KeyboardInterrupt:
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()